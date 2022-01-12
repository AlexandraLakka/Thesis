import os
import csv
import pandas as pd
from openpyxl import load_workbook

# https://stackoverflow.com/questions/12523586/python-format-size-application-converting-b-to-kb-mb-gb-tb/52379087
def humanbytes(B):
    """Return the given bytes as a human friendly KB, MB, GB, or TB string."""
    B = float(B)
    KB = float(1024)
    MB = float(KB ** 2) # 1,048,576
    GB = float(KB ** 3) # 1,073,741,824
    TB = float(KB ** 4) # 1,099,511,627,776

    if B < KB:
        return '{0} {1}'.format(B,'Bytes' if 0 == B > 1 else 'Byte')
    elif KB <= B < MB:
        return float(format(B / KB))
    elif MB <= B < GB:
        #print("MB\n")
        return float(format(B / MB))
    elif GB <= B < TB:
        return float(format(B / GB))
    elif TB <= B:
        return float(format(B / TB))

def write_metrics(virt_mem, res_mem, cpu_sec_tot, function, input_1, input_2):
   cpu = (f"./mem_and_cpu_values/{function}/cpu")
   r_mem = (f"./mem_and_cpu_values/{function}/res_mem")
   v_mem = (f"./mem_and_cpu_values/{function}/virt_mem")
   '''
   with open(os.path.join(v_mem, f"virtual_memory_{input}"), 'w') as myfile:
      for x in virt_mem:
         myfile.write("{}\n".format(round(humanbytes(x), 2)))

   with open(os.path.join(r_mem, f"resident_memory_{input}"), 'w') as myfile:
      for x in res_mem:
         myfile.write("{}\n".format(round(humanbytes(x), 2)))

   with open(os.path.join(cpu, f"cpu_total_time_{input}"), 'w') as myfile:
      for x in cpu_sec_tot:
         myfile.write("{}\n".format(round(x, 2)))
   '''
   with open(os.path.join(v_mem, f"virtual_memory_{input_1}_{input_2}"), 'w') as myfile:
      for x in virt_mem:
         myfile.write("{}\n".format(round(humanbytes(x), 2)))

   with open(os.path.join(r_mem, f"resident_memory_{input_1}_{input_2}"), 'w') as myfile:
      for x in res_mem:
         myfile.write("{}\n".format(round(humanbytes(x), 2)))

   with open(os.path.join(cpu, f"cpu_total_time_{input_1}_{input_2}"), 'w') as myfile:
      for x in cpu_sec_tot:
         myfile.write("{}\n".format(round(x, 2)))

def write_to_excel(virtual, resident, cpu, function, input_1, input_2, input_bytes, size, exec_time):
   # https://stackoverflow.com/questions/47737220/append-dataframe-to-excel-with-pandas
   data = [[2048, function, 2, 'list of integers', input_1, input_2, input_bytes, 'integer', size, virtual[0], virtual[1], virtual[2],
            resident[0], resident[1], resident[2], cpu[0], cpu[1], cpu[2], exec_time]]
   df = pd.DataFrame(data)
   print(df)
   
   excel = f"./{function}.xlsx"
   book = load_workbook(excel)
   writer = pd.ExcelWriter(excel, engine='openpyxl')
   writer.book = book
   writer.sheets = {ws.title: ws for ws in book.worksheets}

   for sheetname in writer.sheets:
      df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

   writer.save()
   
def virtual_memory(virt_mem):
   virt_mem_avg = 0
   sum = 0

   for x in virt_mem:
      a = round(humanbytes(x), 2)
      #print(a)
      sum += a
      #print(sum)

   virt_mem_max = round(humanbytes(max(virt_mem)), 2)
   virt_mem_min = round(humanbytes(min(virt_mem)), 2)
   virt_mem_avg = (sum / len(virt_mem))
   virt_mem_avg = round(virt_mem_avg, 2)

   print(f"Max virt_mem: {virt_mem_max}")
   print(f"Min virt_mem: {virt_mem_min}")
   print(f"Average virt_mem: {virt_mem_avg}")
   print("\n")

   virtual = [virt_mem_max, virt_mem_min, virt_mem_avg]
   return virtual

def resident_memory(res_mem):
   res_mem_avg = 0
   sum = 0

   for x in res_mem:
      a = round(humanbytes(x), 2)
      #print(a)
      sum += a

   res_mem_max = round(humanbytes(max(res_mem)), 2)
   res_mem_min = round(humanbytes(min(res_mem)), 2)
   res_mem_avg = (sum / len(res_mem))
   res_mem_avg = round(res_mem_avg, 2)

   print(f"Max res_mem: {res_mem_max}")
   print(f"Min res_mem: {res_mem_min}")
   print(f"Average res_mem: {res_mem_avg}")
   print("\n")

   resident = [res_mem_max, res_mem_min, res_mem_avg]

   return resident

def cpu_total_time(cpu_sec_tot):
   cpu_sec_tot_avg = 0
   sum = 0

   for x in cpu_sec_tot:
      #print(x)
      sum += x
      # print(cpu_sec_tot_avg)

   cpu_sec_tot_max = round(max(cpu_sec_tot), 2)
   cpu_sec_tot_min = round(min(cpu_sec_tot), 2)
   cpu_sec_tot_avg = (sum / len(cpu_sec_tot))
   cpu_sec_tot_avg = round(cpu_sec_tot_avg, 2)

   print(f"Max cpu_sec_tot: {cpu_sec_tot_max}")
   print(f"Min cpu_sec_tot: {cpu_sec_tot_min}")
   print(f"Average cpu_sec_tot: {cpu_sec_tot_avg}")

   cpu = [cpu_sec_tot_max, cpu_sec_tot_min, cpu_sec_tot_avg]

   return cpu

def main():
   virt_mem = []
   res_mem = []
   cpu_sec_tot = []
   # https://stackoverflow.com/questions/33159106/sort-filenames-in-directory-in-ascending-order
   dir = ("./files")
   file_list = os.listdir(dir)
   file_list = sorted(file_list,key=lambda x: int(os.path.splitext(x)[0]))

   function = 'cholesky_matrix_decomposition'
   input_1 = '580x580'
   input_2 = '-'
   input_bytes = 2691320
   size = len(file_list)
   exec_time = 15.78

   print(f"Batch size: {size}")
   print("\n")

   for filename in file_list:
      with open(os.path.join(dir, filename), 'r') as f:
            lines = f.readlines()
            for line in lines:
               if line.find("process_virtual_memory_bytes") != -1 and line.startswith("#") == False:
                  res = line.partition(" ")[2]
                  res = res.rstrip("\n")
                  virt_mem.append(float(res))
               elif line.find("process_resident_memory_bytes") != -1 and line.startswith("#") == False:
                  res = line.partition(" ")[2]
                  res = res.rstrip("\n")
                  res_mem.append(float(res))
               elif line.find("process_cpu_seconds_total") != -1 and line.startswith("#") == False:
                  res = line.partition(" ")[2]
                  res = res.rstrip("\n")
                  cpu_sec_tot.append(float(res))
               else:
                  continue

   write_metrics(virt_mem, res_mem, cpu_sec_tot, function, input_1, input_2)

   virtual = virtual_memory(virt_mem)
   resident = resident_memory(res_mem)
   cpu = cpu_total_time(cpu_sec_tot)

   write_to_excel(virtual, resident, cpu, function, input_1, input_2, input_bytes, size, exec_time)


if __name__ == "__main__":
   main()
